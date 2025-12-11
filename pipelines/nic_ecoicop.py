# nic_ecoicop.py
"""
NIC ECOICOP Pipeline (Simplified)
=================================
Downloads monthly consumer price indices (NIC) by ECOICOP classification (5 digits)
from ISTAT SDMX API.

Source: ISTAT - Prezzi al consumo per l'intera collettività (NIC)
Dataflow: 167_744_DF_DCSP_NIC1B2015_4
Territory: IT (Italy national)
Classification: ECOICOP 5 digits (all codes via empty string)

Output: NIC_ECOICOP_LATEST.xlsx

Versioning: DateDownload-based (monthly archiving)

This simplified version uses empty string to download ALL ECOICOP codes
in a single request, avoiding rate limiting issues.
"""

import io
import requests
import pandas as pd
import numpy as np
import xml.etree.ElementTree as ET
from datetime import datetime
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time


# =============================================================================
# CONFIGURATION
# =============================================================================

OUTPUT_FILENAME = "NIC_ECOICOP_LATEST.xlsx"

# Output folder ID (subfolder in DATABASE3)
# TODO: Replace with actual Google Drive folder ID for Dati_mensili
OUTPUT_FOLDER_ID = "1Gt7XvvNrnFgpkBEC38lwUXtckurtAK29"

# API URLs
DATAFLOW_ID = "167_744_DF_DCSP_NIC1B2015_4"
STRUCTURE_URL = f"https://esploradati.istat.it/SDMXWS/rest/dataflow/IT1/{DATAFLOW_ID}/1.0/?detail=Full&references=Descendants"

# Data URL - empty string for ECOICOP codes downloads ALL codes
DATA_URL = f"https://esploradati.istat.it/SDMXWS/rest/data/IT1,{DATAFLOW_ID},1.0/M.IT.39.4./ALL/"

# Source path for metadata
SOURCE_PATH = "PRICES / CONSUMER PRICES FOR THE WHOLE NATION / NIC - monthly from 2016 (base 2015) / ECOICOP CLASSIFICATION (5 DIGITS)"
SOURCE_PATH_IT = "PREZZI / PREZZI AL CONSUMO PER L'INTERA COLLETTIVITA / Nic - mensili dal 2016 (base 2015) / CLASSIFICAZIONE ECOICOP (5 CIFRE)"

# Query parameters
START_PERIOD = "2016-01-01"
END_PERIOD = ""

REQUEST_TIMEOUT = 300  # seconds

VERBOSE = True


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def log(msg):
    """Log message with timestamp."""
    if VERBOSE:
        print(f"[NIC_ECOICOP] {msg}")


def get_hierarchy_level(code: str) -> int:
    """Determine ECOICOP hierarchy level from code."""
    if code in ['00', '00ST', 'OR0']:
        return 0
    stripped = code.lstrip('0')
    return len(stripped) if stripped else 1


def fetch_codelist_names() -> dict:
    """
    Fetch ECOICOP code names from ISTAT structure API.
    Returns dict: {code: name_en}
    """
    log("Fetching ECOICOP code labels from structure API...")
    
    try:
        response = requests.get(STRUCTURE_URL, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
    except requests.exceptions.Timeout:
        log("ERROR: Timeout fetching structure - ISTAT server may be slow")
        return {}
    except requests.exceptions.RequestException as e:
        log(f"ERROR: Failed to fetch structure: {e}")
        return {}
    
    try:
        root = ET.fromstring(response.content)
    except ET.ParseError as e:
        log(f"ERROR: Failed to parse structure XML: {e}")
        return {}
    
    names = {}
    
    # Find COICOP codelist
    for cl in root.findall('.//{http://www.sdmx.org/resources/sdmxml/schemas/v2_1/structure}Codelist'):
        cl_id = cl.get('id', '')
        if 'COICOP' in cl_id.upper():
            for code in cl.findall('.//{http://www.sdmx.org/resources/sdmxml/schemas/v2_1/structure}Code'):
                code_id = code.get('id')
                name_en = code_id  # fallback
                for n in code.findall('.//{http://www.sdmx.org/resources/sdmxml/schemas/v2_1/common}Name'):
                    lang = n.get('{http://www.w3.org/XML/1998/namespace}lang')
                    if lang == 'en':
                        name_en = n.text
                        break
                    elif lang == 'it' and name_en == code_id:
                        name_en = n.text
                names[code_id] = name_en
    
    log(f"Loaded {len(names)} ECOICOP code labels")
    return names


def download_nic_data() -> tuple:
    """
    Download NIC ECOICOP data from ISTAT API in a single request.
    Returns tuple of (data: dict, periods: list, error: str or None)
    """
    log("Downloading data from ISTAT (single request with empty string)...")
    log(f"URL: {DATA_URL}")
    
    params = {
        "detail": "full",
        "startPeriod": START_PERIOD,
        "endPeriod": END_PERIOD,
        "dimensionAtObservation": "TIME_PERIOD"
    }
    headers = {"Accept": "application/vnd.sdmx.genericdata+xml;version=2.1"}
    
    ns = {
        'generic': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/data/generic',
        'message': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/message'
    }
    
    try:
        start_time = time.time()
        response = requests.get(DATA_URL, params=params, headers=headers, timeout=REQUEST_TIMEOUT)
        download_time = time.time() - start_time
        
        log(f"Download completed: {len(response.content)/1024/1024:.2f} MB in {download_time:.1f}s")
        
        if response.status_code != 200:
            return {}, [], f"HTTP {response.status_code}"
            
    except requests.exceptions.Timeout:
        return {}, [], "Request timeout"
    except requests.exceptions.RequestException as e:
        return {}, [], str(e)
    
    # Parse XML
    try:
        root = ET.fromstring(response.content)
    except ET.ParseError as e:
        return {}, [], f"XML parse error: {e}"
    
    dataset = root.find('.//message:DataSet', ns)
    if dataset is None:
        return {}, [], "No DataSet found in response"
    
    series_list = dataset.findall('.//generic:Series', ns)
    log(f"Found {len(series_list)} series")
    
    # Extract data
    data = defaultdict(dict)
    periods = set()
    
    for series in series_list:
        series_key = series.find('.//generic:SeriesKey', ns)
        if series_key is None:
            continue
        
        ecoicop_code = None
        for v in series_key.findall('.//generic:Value', ns):
            if v.get('id') == 'E_COICOP_REV_ISTAT':
                ecoicop_code = v.get('value')
                break
        
        if ecoicop_code is None:
            continue
        
        for obs in series.findall('.//generic:Obs', ns):
            obs_dim = obs.find('.//generic:ObsDimension', ns)
            obs_value = obs.find('.//generic:ObsValue', ns)
            
            if obs_dim is not None and obs_value is not None:
                period = obs_dim.get('value', '').replace('-', 'M')
                try:
                    data[ecoicop_code][period] = float(obs_value.get('value'))
                    periods.add(period)
                except (ValueError, TypeError):
                    pass
    
    log(f"Extracted {len(data)} products, {len(periods)} periods")
    return dict(data), sorted(periods), None


def create_excel_file(data: dict, periods: list, code_names: dict, error: str = None) -> io.BytesIO:
    """
    Create Excel file with data and metadata sheets.
    """
    log("Creating Excel file...")
    
    # Create DataFrame
    rows = []
    for code in sorted(data.keys(), key=lambda x: (get_hierarchy_level(x), x)):
        row = {
            'CODE': code,
            'NAME': code_names.get(code, code),
            'LEVEL': get_hierarchy_level(code)
        }
        for period in periods:
            row[period] = data[code].get(period)
        rows.append(row)
    
    df = pd.DataFrame(rows)
    log(f"DataFrame created: {len(df)} rows x {len(df.columns)} columns")
    
    # Create Excel
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Data sheet
        df.to_excel(writer, sheet_name='Data', index=False)
        
        # Format data sheet
        ws = writer.sheets['Data']
        for col_idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            if col in ['CODE', 'NAME']:
                ws.column_dimensions[col_letter].width = 50 if col == 'NAME' else 12
            elif col == 'LEVEL':
                ws.column_dimensions[col_letter].width = 8
            else:
                ws.column_dimensions[col_letter].width = 10
        
        # Metadata sheet
        now = datetime.now()
        metadata = {
            'Field': [
                'edition',
                'edition_type',
                'download_date',
                'source_path',
                'source_path_it',
                'dataflow_id',
                'dataflow_url',
                'measure',
                'measure_code',
                'frequency',
                'frequency_code',
                'base_year',
                'territory',
                'start_period',
                'end_period',
                'n_products',
                'n_periods',
                'errors'
            ],
            'Value': [
                '',
                'DateDownload',
                now.strftime('%Y-%m-%d %H:%M:%S'),
                SOURCE_PATH,
                SOURCE_PATH_IT,
                DATAFLOW_ID,
                STRUCTURE_URL.split('?')[0],
                'Index numbers',
                '4',
                'Monthly',
                'M',
                '2015',
                'IT (Italy)',
                periods[0] if periods else '',
                periods[-1] if periods else '',
                len(df),
                len(periods),
                error if error else 'None'
            ]
        }
        meta_df = pd.DataFrame(metadata)
        meta_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        # Format metadata sheet
        ws_meta = writer.sheets['Metadata']
        ws_meta.column_dimensions['A'].width = 20
        ws_meta.column_dimensions['B'].width = 80
        for row in ws_meta.iter_rows(min_row=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
    
    buffer.seek(0)
    log("Excel file created successfully")
    return buffer


# =============================================================================
# MAIN PIPELINE FUNCTION
# =============================================================================

def run_pipeline() -> dict:
    """
    Main pipeline function.
    Returns dict with status, buffer, and metadata.
    """
    log(f"Pipeline started at {datetime.now().isoformat()}")
    start_time = time.time()
    
    result = {
        'status': 'error',
        'message': '',
        'buffer': None,
        'filename': OUTPUT_FILENAME,
        'metadata': {}
    }
    
    try:
        # 1. Fetch code labels
        code_names = fetch_codelist_names()
        if not code_names:
            log("WARNING: Could not fetch code labels, using codes as names")
        
        # 2. Download data (single request)
        data, periods, error = download_nic_data()
        
        if not data:
            result['message'] = f'Download failed: {error}'
            log(f"ERROR: {result['message']}")
            return result
        
        # 3. Create Excel file
        buffer = create_excel_file(data, periods, code_names, error)
        
        # 4. Prepare result
        elapsed = time.time() - start_time
        
        result['status'] = 'success'
        result['message'] = f'Downloaded {len(data)} products, {len(periods)} periods in {elapsed:.1f}s'
        result['buffer'] = buffer
        result['folder_id'] = OUTPUT_FOLDER_ID
        result['metadata'] = {
            'n_products': len(data),
            'n_periods': len(periods),
            'period_range': f"{periods[0]} → {periods[-1]}" if periods else '',
            'elapsed_seconds': round(elapsed, 1)
        }
        
        log(f"Pipeline completed successfully in {elapsed:.1f}s")
        
    except Exception as e:
        result['message'] = f'Pipeline error: {str(e)}'
        log(f"ERROR: {result['message']}")
        import traceback
        log(traceback.format_exc())
    
    return result


# =============================================================================
# VERSION CONTROL
# =============================================================================

def get_version_info() -> dict:
    """Return version info for the pipeline."""
    now = datetime.now()
    return {
        'version_type': 'DateDownload',
        'version_value': now.strftime('%YM%m'),
        'check_field': None,
        'archive_name': f"NIC_ECOICOP_{now.strftime('%YM%m')}_DateDownload.xlsx"
    }


if __name__ == '__main__':
    result = run_pipeline()
    print(f"\nResult: {result['status']}")
    print(f"Message: {result['message']}")
    if result['metadata']:
        print(f"Metadata: {result['metadata']}")