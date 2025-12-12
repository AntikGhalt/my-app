# istat_consumi_famiglie.py
"""
ISTAT Consumi Famiglie Pipeline
================================
Downloads quarterly household consumption expenditure data from ISTAT SDMX API
and processes it into an Excel file.

Source: ISTAT - Conti economici trimestrali
Dataflow: 163_1226_DF_DCCN_QNA1_3
COICOP: CP01_13 (totale spesa consumi)

Dimensions:
- Aggregato: P31_D_W0_S14 (Territorio+Estero residenti)
             P31_D_W2_S14 (Territorio residenti+non residenti)
- Valutazione: L_2020 (Valori concatenati 2020), V (Prezzi correnti)
- Adjustment: N (Dati grezzi), Y (Dati destagionalizzati)

Output Structure:
- 3 Excel sheets: Metadati, Dati_Grezzi, Dati_Destagionalizzati
- 4 data columns: 2 aggregates × 2 valuations
- Combined column names: "Aggregato - Valutazione"

Versioning: Edition-based

Output: Consumi_famiglie_LATEST.xlsx
Output Folder: Dati_trimestrali (subfolder of DATABASE3)
"""

import io
import requests
import pandas as pd
import numpy as np
import xml.etree.ElementTree as ET
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# =============================================================================
# CONFIGURATION
# =============================================================================

OUTPUT_FILENAME = "Consumi_famiglie_LATEST.xlsx"

# IMPORTANT: Output to Dati_trimestrali subfolder
# TODO: Replace with actual folder ID from Google Drive
OUTPUT_FOLDER_ID = "1GoRcfLt-k3ZAg-j2S1b4k_0ADKy5eYTZ"

# ISTAT API Configuration
BASE_URL = "https://esploradati.istat.it/SDMXWS/rest/data/"
DATAFLOW = "163_1226_DF_DCCN_QNA1_3"

# Query parameters
AGGREGATES = ['P31_D_W0_S14', 'P31_D_W2_S14']
VALUATIONS = ['L_2020', 'V']
ADJUSTMENTS = ['N', 'Y']
COICOP_FILTER = 'CP01_13'  # totale

START_PERIOD = '1775-01-01'
END_PERIOD = ''  # Empty = all available data (no end limit)

# Options
USE_LATEST_EDITION = True
DEFAULT_EDITION = '2025M11'
VERBOSE = True

# Human-readable names
AGGREGATE_NAMES = {
    'P31_D_W0_S14': 'Territorio+Estero (residenti)',
    'P31_D_W2_S14': 'Territorio (residenti+non residenti)'
}

VALUATION_NAMES = {
    'L_2020': 'Valori concatenati 2020',
    'V': 'Prezzi correnti'
}

ADJUSTMENT_NAMES = {
    'N': 'Dati grezzi',
    'Y': 'Dati destagionalizzati'
}

SHEET_NAMES = {
    'N': 'Dati_Grezzi',
    'Y': 'Dati_Destagionalizzati'
}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _log(msg):
    """Print messages only if VERBOSE is True."""
    if VERBOSE:
        print(f"[Consumi_famiglie] {msg}")


def find_latest_edition(max_months_back: int = 24) -> str:
    """
    Find the most recent available edition by searching backwards.
    """
    start_date = datetime.now()
    
    _log("Searching for latest edition...")
    
    headers = {
        "Accept": "application/vnd.sdmx.genericdata+xml;version=2.1"
    }
    
    for i in range(max_months_back + 1):
        test_date = start_date - relativedelta(months=i)
        edition = test_date.strftime('%YM%m')
        
        # Test URL with CP01_13 filter
        url = f"{BASE_URL}IT1,{DATAFLOW},1.0/Q.IT.{AGGREGATES[0]}...{COICOP_FILTER}.{VALUATIONS[0]}.N..{edition}/ALL/"
        
        params = {
            "detail": "full",
            "dimensionAtObservation": "TIME_PERIOD",
            "startPeriod": "2023-01-01",
            "endPeriod": "2023-12-31",
        }
        
        _log(f"  Testing edition: {edition}...")
        
        try:
            response = requests.get(url, params=params, headers=headers, timeout=60)
        except requests.RequestException as e:
            _log(f"    Network error: {e}")
            continue
        
        if response.status_code != 200:
            _log(f"    HTTP {response.status_code}")
            continue
        
        try:
            root = ET.fromstring(response.content)
        except ET.ParseError as e:
            _log(f"    XML error: {e}")
            continue
        
        ns = {
            'msg': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/message',
            'gen': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/data/generic'
        }
        
        dataset = root.find('.//msg:DataSet', ns)
        if dataset is None:
            dataset = root.find('.//gen:DataSet', ns)
        if dataset is None:
            _log("    DataSet not found")
            continue
        
        series_list = dataset.findall('.//gen:Series', ns)
        if not series_list:
            series_list = dataset.findall('.//Series')
        
        if series_list:
            _log(f"  ✅ Found valid edition: {edition}")
            return edition
        
        _log("    No series found")
    
    raise RuntimeError(f"No valid edition found in the last {max_months_back} months.")


def download_istat_data(edition: str) -> tuple[pd.DataFrame, str]:
    """
    Download ISTAT consumption data.
    
    Returns:
        Tuple of (DataFrame, edition_used)
    """
    effective_edition = edition
    
    if USE_LATEST_EDITION:
        try:
            effective_edition = find_latest_edition()
        except RuntimeError as e:
            _log(f"⚠️ Auto edition search failed: {e}")
            _log(f"   Using default edition: {edition}")
            effective_edition = edition
    
    _log(f"Downloading ISTAT data - dataflow {DATAFLOW}")
    _log(f"Edition: {effective_edition}")
    
    # Build query
    agg_str = '+'.join(AGGREGATES)
    val_str = '+'.join(VALUATIONS)
    adj_str = '+'.join(ADJUSTMENTS)
    
    url = f"{BASE_URL}IT1,{DATAFLOW},1.0/Q.IT.{agg_str}...{COICOP_FILTER}.{val_str}.{adj_str}..{effective_edition}/ALL/"
    
    params = {
        "detail": "full",
        "dimensionAtObservation": "TIME_PERIOD",
        "startPeriod": START_PERIOD,
        "endPeriod": END_PERIOD
    }
    
    headers = {
        "Accept": "application/vnd.sdmx.genericdata+xml;version=2.1"
    }
    
    _log(f"URL: {url}")
    
    try:
        response = requests.get(url, params=params, headers=headers, timeout=120)
        response.raise_for_status()
    except requests.RequestException as e:
        _log(f"❌ Download error: {e}")
        return None, None
    
    size_mb = len(response.content) / (1024 * 1024)
    _log(f"Downloaded: {size_mb:.2f} MB")
    
    # Parse XML
    try:
        root = ET.fromstring(response.content)
    except ET.ParseError as e:
        _log(f"❌ XML parsing error: {e}")
        return None, None
    
    ns = {
        'msg': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/message',
        'gen': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/data/generic'
    }
    
    dataset = root.find('.//msg:DataSet', ns)
    if dataset is None:
        dataset = root.find('.//gen:DataSet', ns)
    
    if dataset is None:
        _log("❌ DataSet not found in XML.")
        return None, None
    
    series_list = dataset.findall('.//gen:Series', ns)
    if not series_list:
        series_list = dataset.findall('.//Series')
    
    if not series_list:
        _log("❌ No series found in DataSet.")
        return None, None
    
    _log(f"Series found: {len(series_list)}")
    
    # Extract data
    data_records = []
    
    for series in series_list:
        series_attrs = {}
        
        series_key = series.find('.//gen:SeriesKey', ns)
        if series_key is None:
            series_key = series.find('.//SeriesKey')
        
        if series_key is not None:
            values = series_key.findall('.//gen:Value', ns)
            if not values:
                values = series_key.findall('.//Value')
            for value in values:
                dim_id = value.get('id')
                dim_value = value.get('value')
                if dim_id and dim_value:
                    series_attrs[dim_id] = dim_value
        
        obs_list = series.findall('.//gen:Obs', ns)
        if not obs_list:
            obs_list = series.findall('.//Obs')
        
        for obs in obs_list:
            record = series_attrs.copy()
            
            obs_dim = obs.find('.//gen:ObsDimension', ns)
            if obs_dim is None:
                obs_dim = obs.find('.//ObsDimension')
            period = obs_dim.get('value', '') if obs_dim is not None else ''
            record['TIME_PERIOD'] = period
            
            obs_value = obs.find('.//gen:ObsValue', ns)
            if obs_value is None:
                obs_value = obs.find('.//ObsValue')
            
            if obs_value is not None:
                v = obs_value.get('value', None)
                try:
                    record['VALUE'] = float(v) if v is not None else np.nan
                except (TypeError, ValueError):
                    record['VALUE'] = np.nan
            else:
                record['VALUE'] = np.nan
            
            data_records.append(record)
    
    df = pd.DataFrame(data_records)
    _log(f"Total observations: {df.shape[0]}")
    
    return df, effective_edition


def process_data(df: pd.DataFrame) -> dict:
    """
    Process raw data into structured format.
    
    Returns:
        Dictionary with keys 'N' and 'Y' (adjustments), each containing a DataFrame
    """
    if df is None or df.empty:
        return None
    
    _log("Processing data...")
    
    result = {}
    
    for adj in ADJUSTMENTS:
        _log(f"  Processing: {ADJUSTMENT_NAMES[adj]} ({adj})")
        
        df_adj = df[df['ADJUSTMENT'] == adj].copy()
        
        if df_adj.empty:
            _log(f"    No data for adjustment {adj}")
            continue
        
        records = []
        
        for agg in AGGREGATES:
            for val in VALUATIONS:
                mask = (df_adj['DATA_TYPE_AGGR'] == agg) & (df_adj['VALUATION'] == val)
                subset = df_adj[mask]
                
                if subset.empty:
                    continue
                
                # Create combined name: "Aggregato - Valutazione"
                var_name = f"{AGGREGATE_NAMES.get(agg, agg)} - {VALUATION_NAMES.get(val, val)}"
                
                for _, row in subset.iterrows():
                    records.append({
                        'TIME_PERIOD': row['TIME_PERIOD'],
                        'VARIABLE': var_name,
                        'VALUE': row['VALUE']
                    })
        
        if not records:
            continue
        
        df_long = pd.DataFrame(records)
        
        # Handle duplicates
        df_long = df_long.drop_duplicates(subset=['TIME_PERIOD', 'VARIABLE'], keep='first')
        
        # Pivot to wide format
        df_wide = df_long.pivot(index='TIME_PERIOD', columns='VARIABLE', values='VALUE')
        df_wide = df_wide.sort_index()
        
        result[adj] = df_wide
        _log(f"    Created: {df_wide.shape[0]} rows × {df_wide.shape[1]} columns")
    
    return result


def create_excel_file(data_dict: dict, edition: str) -> io.BytesIO:
    """
    Create Excel file with metadata and data sheets.
    """
    _log("Creating Excel file...")
    
    buffer = io.BytesIO()
    
    # Get period range
    period_min = None
    period_max = None
    for df_d in data_dict.values():
        if df_d is not None and not df_d.empty:
            if period_min is None:
                period_min = df_d.index.min()
                period_max = df_d.index.max()
            else:
                period_min = min(period_min, df_d.index.min())
                period_max = max(period_max, df_d.index.max())
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        
        # Sheet 1: Metadati
        meta_rows = [
            ('edition', edition),
            ('edition_type', 'Edition'),
            ('download_date', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ('dataflow', DATAFLOW),
            ('description', 'Spesa per consumi finali delle famiglie'),
            ('coicop_filter', f'{COICOP_FILTER} (totale)'),
            ('period_min', str(period_min)),
            ('period_max', str(period_max)),
            ('n_variables', len(AGGREGATES) * len(VALUATIONS)),
        ]
        df_meta = pd.DataFrame(meta_rows, columns=['chiave', 'valore'])
        df_meta.to_excel(writer, sheet_name='Metadati', index=False)
        
        # Variable metadata
        var_meta_rows = []
        for agg in AGGREGATES:
            for val in VALUATIONS:
                var_meta_rows.append({
                    'variable_name': f"{AGGREGATE_NAMES.get(agg, agg)} - {VALUATION_NAMES.get(val, val)}",
                    'aggregate_code': agg,
                    'aggregate_name': AGGREGATE_NAMES.get(agg, agg),
                    'valuation_code': val,
                    'valuation_name': VALUATION_NAMES.get(val, val)
                })
        df_var_meta = pd.DataFrame(var_meta_rows)
        df_var_meta.to_excel(writer, sheet_name='Metadati', index=False, startrow=len(meta_rows)+3)
        
        # Format metadata sheet
        ws_meta = writer.sheets['Metadati']
        ws_meta.column_dimensions['A'].width = 20
        ws_meta.column_dimensions['B'].width = 60
        for col_idx in range(3, 6):
            col_letter = get_column_letter(col_idx)
            ws_meta.column_dimensions[col_letter].width = 40
        
        # Data sheets
        for adj, sheet_name in SHEET_NAMES.items():
            if adj not in data_dict or data_dict[adj] is None:
                continue
            
            df_data = data_dict[adj].copy()
            
            # Add temporal columns
            per_idx = pd.PeriodIndex(df_data.index, freq='Q')
            df_data.insert(0, 'PERIOD', per_idx.astype(str))
            df_data.insert(1, 'YEAR', per_idx.year)
            df_data.insert(2, 'SEMESTER', np.where(per_idx.quarter <= 2, 'Sem1', 'Sem2'))
            df_data.insert(3, 'QUARTER', [f'Q{q}' for q in per_idx.quarter])
            
            df_data.reset_index(drop=True, inplace=True)
            df_data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format data sheet
            ws = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(df_data.columns, start=1):
                col_letter = get_column_letter(col_idx)
                if col_name in ('PERIOD', 'YEAR', 'SEMESTER', 'QUARTER'):
                    ws.column_dimensions[col_letter].width = 10
                else:
                    ws.column_dimensions[col_letter].width = 40
            
            for cell in ws[1]:
                cell.alignment = Alignment(wrap_text=True)
    
    _log("Excel file created successfully")
    return buffer


# =============================================================================
# MAIN PIPELINE FUNCTION
# =============================================================================

def run_pipeline() -> dict:
    """
    Execute the complete pipeline: download, process, create Excel.
    
    Returns:
        Dictionary with:
        - status: 'success' or 'error'
        - buffer: BytesIO with Excel file (if success)
        - filename: Output filename
        - edition: Data edition
        - folder_id: Target folder ID for upload
        - metadata: Additional information
    """
    _log(f"Pipeline started at {datetime.now().isoformat()}")
    
    try:
        # 1. Download raw data
        _log("Downloading ISTAT data...")
        df, used_edition = download_istat_data(DEFAULT_EDITION)
        
        if df is None or df.empty:
            return {
                'status': 'error',
                'message': 'Download failed, no data received'
            }
        
        _log(f"Downloaded {len(df)} observations, edition: {used_edition}")
        
        # 2. Process data
        _log("Processing data...")
        data_dict = process_data(df)
        
        if not data_dict:
            return {
                'status': 'error',
                'message': 'No data extracted after processing'
            }
        
        # 3. Create Excel file
        buffer = create_excel_file(data_dict, used_edition)
        
        # Calculate stats
        total_obs = sum(d.shape[0] for d in data_dict.values() if d is not None)
        n_sheets = len(data_dict)
        
        _log(f"Pipeline completed successfully")
        
        return {
            'status': 'success',
            'buffer': buffer,
            'filename': OUTPUT_FILENAME,
            'edition': used_edition,
            'folder_id': OUTPUT_FOLDER_ID,  # Custom output folder
            'n_observations': total_obs,
            'n_sheets': n_sheets,
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'status': 'error',
            'message': str(e)
        }


# =============================================================================
# STANDALONE EXECUTION (for testing)
# =============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("ISTAT Consumi Famiglie Pipeline - Test Run")
    print("=" * 70)
    
    result = run_pipeline()
    
    print(f"\nResult: {result['status']}")
    
    if result['status'] == 'success':
        print(f"Edition: {result['edition']}")
        print(f"Observations: {result['n_observations']}")
        print(f"Sheets: {result['n_sheets']}")
        print(f"Target folder: {result['folder_id']}")
        
        # Save to file for inspection
        with open('Consumi_famiglie_TEST.xlsx', 'wb') as f:
            f.write(result['buffer'].getvalue())
        print("\nSaved to: Consumi_famiglie_TEST.xlsx")
    else:
        print(f"Error: {result.get('message', 'Unknown error')}")