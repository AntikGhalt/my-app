# istat_reddito_famiglie.py

"""
ISTAT Reddito Disponibile Famiglie Pipeline
============================================
Downloads quarterly household disposable income data from ISTAT SDMX API
and processes it into an Excel file.

Source: ISTAT - Conti economici trimestrali dei settori istituzionali
Dataflow: 162_1064_DF_DCCN_ISTITUZ_QNA1_1
Sector: S14A (Consumer households / Famiglie consumatrici)

Output: Reddito_disponibile_famiglie_LATEST.xlsx
"""

import io
import requests
import pandas as pd
import numpy as np
import xml.etree.ElementTree as ET
from datetime import datetime
from dateutil.relativedelta import relativedelta
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# =============================================================================
# CONFIGURATION
# =============================================================================

# Output filename (fixed name for Tableau)
OUTPUT_FILENAME = "ISTAT_Reddito_disponibile_famiglie_LATEST.xlsx"

# Output folder ID (subfolder in DATABASE3)
# TODO: Replace with actual Google Drive folder ID for Dati_trimestrali
OUTPUT_FOLDER_ID = "1GoRcfLt-k3ZAg-j2S1b4k_0ADKy5eYTZ"

# ISTAT API Configuration
BASE_URL = "https://esploradati.istat.it/SDMXWS/rest/data/"
DATAFLOW = "162_1064_DF_DCCN_ISTITUZ_QNA1_1"

# Query parameters
SECTOR = 'S14A'                     # S14A = Consumer households (Famiglie consumatrici)
START_PERIOD = '1775-07-01'         # All available historical data
END_PERIOD = '2030-07-01'           # Far future to get all data

# Options
USE_LATEST_EDITION = True           # Automatically find most recent edition
DEFAULT_EDITION = '2025M10'         # Fallback edition
NEGATE_IMPIEGO = True               # Make IMPIEGO series negative
VERBOSE = False                     # Debug logging

# Aggregates to download (1 = download, 0 = ignore)
AGGREGATES = {
    # CONTO DELLA ATTRIBUZIONE DEI REDDITI PRIMARI
    'B2A3G_B_W0_X1': 1,   # RISORSA - Risultato lordo di gestione e reddito misto lordo
    'D1_C_W0': 1,         # RISORSA - Redditi da lavoro dipendente
    'D4T_C_W0': 1,        # RISORSA - Redditi da capitale (comprensivi quota famiglie produttrici)
    'D4T_D_W0': 1,        # IMPIEGO - Redditi da capitale (comprensivi quota famiglie produttrici)
    'B5G_B_W0': 1,        # SALDO - Reddito nazionale lordo/saldo dei redditi primari lordo
    
    # CONTO DELLA DISTRIBUZIONE SECONDARIA DEL REDDITO
    'D61_C_W0': 1,        # RISORSA - Contributi sociali netti
    'D62_C_W0': 1,        # RISORSA - Prestazioni sociali diverse dai trasferimenti sociali in natura
    'D7_C_W0': 1,         # RISORSA - Altri trasferimenti correnti
    'D5_D_W0': 1,         # IMPIEGO - Imposte correnti sul reddito, sul patrimonio, ecc.
    'D61_D_W0': 1,        # IMPIEGO - Contributi sociali netti
    'D62_D_W0': 1,        # IMPIEGO - Prestazioni sociali diverse dai trasferimenti sociali in natura
    'D7_D_W0': 1,         # IMPIEGO - Altri trasferimenti correnti
    'B6G_B_W0': 1,        # SALDO - Reddito disponibile lordo
    
    # CONTO DI UTILIZZAZIONE DEL REDDITO DISPONIBILE
    'D8_C_W0': 1,         # RISORSA - Rettifica per variazione dei diritti pensionistici
}

# Descriptive names for aggregates
AGGREGATES_NAMES = {
    'B2A3G_B_W0_X1': 'RISORSA - Risultato lordo di gestione e reddito misto lordo',
    'D1_C_W0': 'RISORSA - Redditi da lavoro dipendente',
    'D4T_C_W0': 'RISORSA - Redditi da capitale (comprensivi quota famiglie produttrici)',
    'D4T_D_W0': 'IMPIEGO - Redditi da capitale (comprensivi quota famiglie produttrici)',
    'B5G_B_W0': 'SALDO - Reddito nazionale lordo/saldo dei redditi primari lordo',
    'D61_C_W0': 'RISORSA - Contributi sociali netti',
    'D62_C_W0': 'RISORSA - Prestazioni sociali diverse dai trasferimenti sociali in natura',
    'D7_C_W0': 'RISORSA - Altri trasferimenti correnti',
    'D5_D_W0': 'IMPIEGO - Imposte correnti sul reddito, sul patrimonio, ecc.',
    'D61_D_W0': 'IMPIEGO - Contributi sociali netti',
    'D62_D_W0': 'IMPIEGO - Prestazioni sociali diverse dai trasferimenti sociali in natura',
    'D7_D_W0': 'IMPIEGO - Altri trasferimenti correnti',
    'B6G_B_W0': 'SALDO - Reddito disponibile lordo',
    'D8_C_W0': 'RISORSA - Rettifica per variazione dei diritti pensionistici',
}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _log(msg):
    """Print messages only if VERBOSE is True."""
    if VERBOSE:
        print(msg)


def find_latest_edition(sector: str, aggregates_str: str, max_months_back: int = 24) -> str:
    """
    Find the most recent available edition by searching backwards.
    
    Args:
        sector: Institutional sector code
        aggregates_str: Plus-separated aggregate codes
        max_months_back: How many months to search backwards
    
    Returns:
        Edition string (e.g., "2025M10")
    
    Raises:
        RuntimeError: If no valid edition found
    """
    start_date = datetime.now()
    
    _log("=" * 70)
    _log("SEARCHING FOR LATEST EDITION")
    _log("=" * 70)

    headers = {
        "Accept": "application/vnd.sdmx.genericdata+xml;version=2.1"
    }

    for i in range(max_months_back + 1):
        test_date = start_date - relativedelta(months=i)
        edition = f"{test_date.year}M{test_date.month}"
        url = f"{BASE_URL}IT1,{DATAFLOW},1.0/Q.IT.{aggregates_str}.{sector}...V.S.N.{edition}/ALL/"
        
        edition_year = test_date.year
        test_year = edition_year - 1
        params = {
            "detail": "full",
            "dimensionAtObservation": "TIME_PERIOD",
            "startPeriod": f"{test_year}-01-01",
            "endPeriod": f"{test_year}-12-31",
        }

        _log(f"Attempt {i+1}: {edition} ...")

        try:
            response = requests.get(url, params=params, headers=headers, timeout=30)
        except requests.RequestException as e:
            _log(f"  Network error ({type(e).__name__}): {e}")
            continue

        if response.status_code != 200:
            _log(f"  HTTP {response.status_code}")
            continue

        try:
            root = ET.fromstring(response.content)
        except ET.ParseError as e:
            _log(f"  XML error: {e}")
            continue

        ns = {
            'msg': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/message',
            'gen': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/data/generic'
        }

        dataset = root.find('.//msg:DataSet', ns)
        if dataset is None:
            dataset = root.find('.//gen:DataSet', ns)
        if dataset is None:
            dataset = root.find('.//DataSet')

        if dataset is None:
            _log("  DataSet not found")
            continue

        series_list = dataset.findall('.//gen:Series', ns)
        if not series_list:
            series_list = dataset.findall('.//Series')

        if series_list:
            _log(f"  ✅ Valid edition: {edition}")
            return edition

        _log("  No series found")

    raise RuntimeError(
        f"No valid edition found in the last {max_months_back} months."
    )


def download_istat_data(aggregates_dict: dict, sector: str, edition: str,
                        start_period: str = None, end_period: str = None,
                        use_latest_edition: bool = True) -> tuple[pd.DataFrame, str]:
    """
    Download ISTAT data for specified aggregates and sector.
    
    Args:
        aggregates_dict: Dictionary of aggregate codes with 1/0 flags
        sector: Institutional sector code
        edition: Edition to use if not searching for latest
        start_period: Start period filter
        end_period: End period filter
        use_latest_edition: Whether to search for latest edition
    
    Returns:
        Tuple of (DataFrame, edition_used)
    """
    active_aggregates = [agg for agg, flag in aggregates_dict.items() if flag == 1]

    if not active_aggregates:
        print("No aggregates selected (all flags = 0).")
        return None, None

    aggregates_str = '+'.join(active_aggregates)

    # Find latest edition if requested
    effective_edition = edition
    if use_latest_edition:
        try:
            effective_edition = find_latest_edition(sector, aggregates_str)
        except RuntimeError as e:
            print(f"⚠️  Auto edition search failed: {e}")
            print(f"   Using specified edition: {edition}")
            effective_edition = edition

    _log(f"Downloading ISTAT data - dataflow {DATAFLOW}")
    _log(f"Sector: {sector}")
    _log(f"Active aggregates: {len(active_aggregates)}")
    _log(f"Edition: {effective_edition}")

    url = f"{BASE_URL}IT1,{DATAFLOW},1.0/Q.IT.{aggregates_str}.{sector}...V.S.N.{effective_edition}/ALL/"

    params = {
        "detail": "full",
        "dimensionAtObservation": "TIME_PERIOD"
    }
    if start_period:
        params["startPeriod"] = start_period
    if end_period:
        params["endPeriod"] = end_period

    headers = {
        "Accept": "application/vnd.sdmx.genericdata+xml;version=2.1"
    }

    try:
        response = requests.get(url, params=params, headers=headers, timeout=120)
        response.raise_for_status()
    except requests.RequestException as e:
        print(f"❌ Download error: {e}")
        return None, None

    size_mb = len(response.content) / (1024 * 1024)
    _log(f"Downloaded: {size_mb:.1f} MB")

    # Parse XML
    try:
        root = ET.fromstring(response.content)
    except ET.ParseError as e:
        print(f"❌ XML parsing error: {e}")
        return None, None

    ns = {
        'msg': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/message',
        'gen': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/data/generic'
    }

    dataset = root.find('.//msg:DataSet', ns)
    if dataset is None:
        dataset = root.find('.//gen:DataSet', ns)
    if dataset is None:
        dataset = root.find('.//DataSet')

    if dataset is None:
        print("❌ DataSet not found in XML.")
        return None, None

    series_list = dataset.findall('.//gen:Series', ns)
    if not series_list:
        series_list = dataset.findall('.//Series')

    if not series_list:
        print("❌ No series found in DataSet.")
        return None, None

    _log(f"Series found: {len(series_list)}")

    # Extract data
    data_dict = defaultdict(list)

    for series in series_list:
        series_attrs = {}

        series_key = series.find('.//gen:SeriesKey', ns)
        if series_key is None:
            series_key = series.find('.//SeriesKey')

        if series_key is not None:
            values = series_key.findall('.//gen:Value', ns)
            if not values:
                values = series_key.findall('.//Value')
            for value in values:
                dim_id = value.get('id')
                dim_value = value.get('value')
                if dim_id and dim_value:
                    series_attrs[dim_id] = dim_value

        obs_list = series.findall('.//gen:Obs', ns)
        if not obs_list:
            obs_list = series.findall('.//Obs')

        for obs in obs_list:
            for key, val in series_attrs.items():
                data_dict[key].append(val)

            obs_dim = obs.find('.//gen:ObsDimension', ns)
            if obs_dim is None:
                obs_dim = obs.find('.//ObsDimension')
            period = obs_dim.get('value', '') if obs_dim is not None else ''
            data_dict['TIME_PERIOD'].append(period)

            obs_value = obs.find('.//gen:ObsValue', ns)
            if obs_value is None:
                obs_value = obs.find('.//ObsValue')

            if obs_value is not None:
                v = obs_value.get('value', None)
                try:
                    data_dict['VALUE'].append(float(v) if v is not None else pd.NA)
                except (TypeError, ValueError):
                    data_dict['VALUE'].append(pd.NA)
            else:
                data_dict['VALUE'].append(pd.NA)

    df = pd.DataFrame(data_dict)
    df['VALUE'] = pd.to_numeric(df['VALUE'], errors='coerce')

    _log(f"Total observations: {df.shape[0]}")
    return df, effective_edition


def extract_series(df: pd.DataFrame, aggregate_code: str, sector: str) -> pd.Series:
    """
    Extract time series for a specific aggregate and sector.
    
    Args:
        df: Raw DataFrame from ISTAT
        aggregate_code: Aggregate code to extract
        sector: Sector code
    
    Returns:
        pandas Series with PeriodIndex
    """
    if df is None or df.empty:
        return None

    required_cols = {'DATA_TYPE_AGGR', 'INSTITUTIONAL_SECTOR', 'TIME_PERIOD', 'VALUE'}
    if not required_cols.issubset(df.columns):
        print("❌ Required columns not found in DataFrame.")
        return None

    mask = (
        (df['DATA_TYPE_AGGR'] == aggregate_code) &
        (df['INSTITUTIONAL_SECTOR'] == sector) &
        (df['VALUE'].notna())
    )
    subset = df[mask]
    if subset.empty:
        return None

    subset = subset.sort_values('TIME_PERIOD')

    try:
        index = pd.PeriodIndex(subset['TIME_PERIOD'], freq='Q')
    except Exception:
        index = pd.to_datetime(subset['TIME_PERIOD'], errors='coerce')

    ts = pd.Series(subset['VALUE'].values, index=index, name=aggregate_code)
    return ts


def classify_aggregate(code: str) -> tuple[str, str, str]:
    """
    Classify an aggregate code and extract clean name.
    
    Args:
        code: Aggregate code
    
    Returns:
        Tuple of (clean_name, flow_direction, raw_label)
        flow_direction is one of: 'RISORSA', 'IMPIEGO', 'SALDO', 'AMMORTAMENTO', None
    """
    label = AGGREGATES_NAMES.get(code, code)
    clean_label = label
    flow = None

    for prefix in ('RISORSA', 'IMPIEGO', 'SALDO'):
        full_prefix = prefix + ' - '
        if label.startswith(full_prefix):
            clean_label = label[len(full_prefix):].strip()
            flow = prefix
            break

    if flow is None:
        lower = label.lower()
        if lower.startswith('risorse') or lower.startswith('risorsa'):
            flow = 'RISORSA'
        elif lower.startswith('impieghi') or lower.startswith('impiego'):
            flow = 'IMPIEGO'
        elif lower.startswith('saldo'):
            flow = 'SALDO'
        elif 'ammortamenti' in lower:
            flow = 'AMMORTAMENTO'

    return clean_label, flow, label


def build_series_metadata(codes: list) -> dict:
    """
    Build metadata dictionary for aggregate codes.
    
    Args:
        codes: List of aggregate codes
    
    Returns:
        Dictionary with metadata for each code
    """
    meta = {}
    for code in codes:
        clean_name, flow, raw_label = classify_aggregate(code)
        meta[code] = {
            "code": code,
            "name": clean_name,
            "raw_label": raw_label,
            "flow_direction": flow
        }
    return meta


def apply_flow_signs(series_dict: dict, series_meta: dict, negate_impiego: bool = True) -> dict:
    """
    Make IMPIEGO series negative if requested.
    
    Args:
        series_dict: Dictionary of series
        series_meta: Metadata dictionary
        negate_impiego: Whether to negate IMPIEGO flows
    
    Returns:
        Adjusted series dictionary
    """
    if not negate_impiego:
        return series_dict

    adjusted = {}
    for code, ts in series_dict.items():
        info = series_meta.get(code)
        if info and info.get("flow_direction") == "IMPIEGO":
            adjusted[code] = -ts
        else:
            adjusted[code] = ts
    return adjusted


# =============================================================================
# MAIN PIPELINE FUNCTION
# =============================================================================

def run_pipeline() -> dict:
    """
    Execute the complete pipeline: download, process, create Excel.
    
    Returns:
        Dictionary with:
        - status: 'success' or 'error'
        - buffer: BytesIO with Excel file (if success)
        - filename: Output filename
        - edition: Data edition
        - folder_id: Target folder for upload
        - metadata: Additional information
    """
    print(f"[Reddito_disponibile_famiglie] Pipeline started at {datetime.now().isoformat()}")
    
    try:
        # 1. Download raw data
        print("[Reddito_disponibile_famiglie] Downloading ISTAT data...")
        df, used_edition = download_istat_data(
            aggregates_dict=AGGREGATES,
            sector=SECTOR,
            edition=DEFAULT_EDITION,
            start_period=START_PERIOD,
            end_period=END_PERIOD,
            use_latest_edition=USE_LATEST_EDITION
        )

        if df is None or df.empty:
            return {
                'status': 'error',
                'message': 'Download failed, no data received'
            }

        print(f"[Reddito_disponibile_famiglie] Downloaded {len(df)} observations, edition: {used_edition}")

        # 2. Extract series for active aggregates
        active_aggregates = [agg for agg, flag in AGGREGATES.items() if flag == 1]
        series_dict = {}

        for agg in active_aggregates:
            ts = extract_series(df, agg, SECTOR)
            if ts is not None:
                series_dict[agg] = ts

        if not series_dict:
            return {
                'status': 'error',
                'message': 'No series extracted for selected aggregates'
            }

        print(f"[Reddito_disponibile_famiglie] Extracted {len(series_dict)} series")

        # 3. Build metadata
        series_meta = build_series_metadata(series_dict.keys())

        # 4. Apply negative sign to IMPIEGO
        series_dict = apply_flow_signs(series_dict, series_meta, negate_impiego=NEGATE_IMPIEGO)

        # 5. Create wide-format DataFrame
        df_output = pd.DataFrame(series_dict)
        df_output = df_output.sort_index()

        # Period index processing
        if isinstance(df_output.index, pd.PeriodIndex):
            per_idx = df_output.index
        else:
            per_idx = df_output.index.to_period('Q')

        period_min = per_idx.min()
        period_max = per_idx.max()

        period_str = per_idx.astype(str)
        year_arr = per_idx.year
        quarter_arr = per_idx.quarter
        semester_arr = np.where(quarter_arr <= 2, "Sem1", "Sem2")
        quarter_label_arr = np.array([f"Q{q}" for q in quarter_arr])

        # Insert temporal columns
        df_output.insert(0, "PERIOD", period_str)
        df_output.insert(1, "YEAR", year_arr)
        df_output.insert(2, "SEMESTER", semester_arr)
        df_output.insert(3, "QUARTER", quarter_label_arr)

        df_output.reset_index(drop=True, inplace=True)

        # 6. Create DataFrame with clean names
        df_output_nomi = df_output.copy()
        meta_cols = ["PERIOD", "YEAR", "SEMESTER", "QUARTER"]
        var_cols = [c for c in df_output.columns if c not in meta_cols]

        rename_map = {code: series_meta[code]["name"] for code in var_cols}
        df_output_nomi.rename(columns=rename_map, inplace=True)

        # 7. Prepare Excel file in memory
        print("[Reddito_disponibile_famiglie] Creating Excel file...")
        
        # Global metadata
        global_meta_rows = [
            ("edition", used_edition),
            ("edition_type", "Edition"),
            ("download_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("sector", SECTOR),
            ("sector_description", "Famiglie consumatrici"),
            ("dataflow", DATAFLOW),
            ("period_min", str(period_min)),
            ("period_max", str(period_max)),
            ("download_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("n_variables", len(var_cols)),
        ]
        df_global_meta = pd.DataFrame(global_meta_rows, columns=["chiave", "valore"])

        # Variable metadata
        df_vars_meta = pd.DataFrame(
            [
                {
                    "code": info["code"],
                    "name": info["name"],
                    "raw_label": info["raw_label"],
                    "flow_direction": info["flow_direction"],
                }
                for code, info in sorted(series_meta.items())
            ]
        )

        # Write to BytesIO buffer
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # Sheet 1: Metadati
            df_global_meta.to_excel(writer, sheet_name="Metadati", index=False)
            startrow = len(df_global_meta) + 2
            df_vars_meta.to_excel(writer, sheet_name="Metadati", index=False, startrow=startrow)

            # Sheet 2: Dati
            df_output_nomi.to_excel(writer, sheet_name="Dati", index=False)

            # Formatting
            ws_data = writer.sheets["Dati"]
            ws_meta = writer.sheets["Metadati"]

            # Format Dati sheet
            for col_idx, col_name in enumerate(df_output_nomi.columns, start=1):
                col_letter = get_column_letter(col_idx)
                if col_name in ("PERIOD", "YEAR", "SEMESTER", "QUARTER"):
                    ws_data.column_dimensions[col_letter].width = 10
                else:
                    ws_data.column_dimensions[col_letter].width = 25

            for cell in ws_data[1]:
                cell.alignment = Alignment(wrap_text=True)

            # Format Metadati sheet
            for col_idx in range(1, 5):
                col_letter = get_column_letter(col_idx)
                ws_meta.column_dimensions[col_letter].width = 30

        print(f"[Reddito_disponibile_famiglie] Pipeline completed successfully")
        
        return {
            'status': 'success',
            'buffer': buffer,
            'filename': OUTPUT_FILENAME,
            'edition': used_edition,
            'folder_id': OUTPUT_FOLDER_ID,
            'n_variables': len(var_cols),
            'n_observations': len(df_output),
            'period_range': f"{period_min} → {period_max}",
            'sector': SECTOR
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'status': 'error',
            'message': str(e)
        }


# =============================================================================
# STANDALONE EXECUTION (for testing)
# =============================================================================

if __name__ == "__main__":
    result = run_pipeline()
    print(f"\nResult: {result['status']}")
    if result['status'] == 'success':
        print(f"Edition: {result['edition']}")
        print(f"Variables: {result['n_variables']}")
        print(f"Observations: {result['n_observations']}")
        print(f"Period: {result['period_range']}")