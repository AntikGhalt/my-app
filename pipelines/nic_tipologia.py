# nic_tipologia.py
"""
NIC Tipologia Pipeline (Simplified)
====================================
Downloads monthly consumer price indices (NIC) by product type and territory
from ISTAT SDMX API.

Source: ISTAT - Prezzi al consumo per l'intera collettività (NIC)
Dataflow: 167_744_DF_DCSP_NIC1B2015_2
Territories: IT + 5 macro-areas (IT, ITC, ITD, ITE, ITF, ITG)
Classification: Product types (all codes via empty string)

Output: NIC_Tipologia_LATEST.xlsx

Versioning: DateDownload-based (monthly archiving)

This simplified version downloads only 6 territories (Italy + 5 macro-areas)
instead of all 132 territories, reducing download from 24 MB to ~1.6 MB.
"""

import io
import requests
import pandas as pd
import numpy as np
import xml.etree.ElementTree as ET
from datetime import datetime
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time


# =============================================================================
# CONFIGURATION
# =============================================================================

OUTPUT_FILENAME = "NIC_Tipologia_LATEST.xlsx"

# Output folder ID (subfolder in DATABASE3)
# TODO: Replace with actual Google Drive folder ID for Dati_mensili
OUTPUT_FOLDER_ID = "1Gt7XvvNrnFgpkBEC38lwUXtckurtAK29"

# API URLs
DATAFLOW_ID = "167_744_DF_DCSP_NIC1B2015_2"
STRUCTURE_URL = f"https://esploradati.istat.it/SDMXWS/rest/dataflow/IT1/{DATAFLOW_ID}/1.0/?detail=Full&references=Descendants"

# Data URL - 6 territories + empty string for product types
# IT=Italia, ITC=Nord-ovest, ITD=Nord-est, ITE=Centro, ITF=Sud, ITG=Isole
TERRITORIES = "IT+ITC+ITD+ITE+ITF+ITG"
DATA_URL = f"https://esploradati.istat.it/SDMXWS/rest/data/IT1,{DATAFLOW_ID},1.0/M.{TERRITORIES}.39.4./ALL/"

# Territory names
TERRITORY_NAMES = {
    'IT': 'Italia',
    'ITC': 'Nord-ovest',
    'ITD': 'Nord-est', 
    'ITE': 'Centro',
    'ITF': 'Sud',
    'ITG': 'Isole'
}

# Source path for metadata
SOURCE_PATH = "PRICES / CONSUMER PRICES FOR THE WHOLE NATION / NIC - monthly from 2016 (base 2015) / PRODUCT TYPE"
SOURCE_PATH_IT = "PREZZI / PREZZI AL CONSUMO PER L'INTERA COLLETTIVITA / Nic - mensili dal 2016 (base 2015) / TIPOLOGIA DI PRODOTTO"

# Query parameters
START_PERIOD = "2016-01-01"
END_PERIOD = ""

REQUEST_TIMEOUT = 300  # seconds

VERBOSE = True


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def log(msg):
    """Log message with timestamp."""
    if VERBOSE:
        print(f"[NIC_Tipologia] {msg}")


def fetch_codelist_names() -> dict:
    """
    Fetch product type code names from ISTAT structure API.
    Returns dict: {code: name_en}
    """
    log("Fetching product type labels from structure API...")
    
    try:
        response = requests.get(STRUCTURE_URL, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
    except requests.exceptions.Timeout:
        log("ERROR: Timeout fetching structure - ISTAT server may be slow")
        return {}
    except requests.exceptions.RequestException as e:
        log(f"ERROR: Failed to fetch structure: {e}")
        return {}
    
    try:
        root = ET.fromstring(response.content)
    except ET.ParseError as e:
        log(f"ERROR: Failed to parse structure XML: {e}")
        return {}
    
    product_names = {}
    
    # Find COICOP codelist for product types
    for cl in root.findall('.//{http://www.sdmx.org/resources/sdmxml/schemas/v2_1/structure}Codelist'):
        cl_id = cl.get('id', '')
        if 'COICOP' in cl_id.upper():
            for code in cl.findall('.//{http://www.sdmx.org/resources/sdmxml/schemas/v2_1/structure}Code'):
                code_id = code.get('id')
                name_en = code_id  # fallback
                for n in code.findall('.//{http://www.sdmx.org/resources/sdmxml/schemas/v2_1/common}Name'):
                    lang = n.get('{http://www.w3.org/XML/1998/namespace}lang')
                    if lang == 'en':
                        name_en = n.text
                        break
                    elif lang == 'it' and name_en == code_id:
                        name_en = n.text
                product_names[code_id] = name_en
    
    log(f"Loaded {len(product_names)} product type labels")
    return product_names


def download_nic_data() -> tuple:
    """
    Download NIC Tipologia data from ISTAT API in a single request.
    Returns tuple of (data: dict, periods: list, error: str or None)
    """
    log("Downloading data from ISTAT (single request)...")
    log(f"URL: {DATA_URL}")
    log(f"Territories: {TERRITORIES}")
    
    params = {
        "detail": "full",
        "startPeriod": START_PERIOD,
        "endPeriod": END_PERIOD,
        "dimensionAtObservation": "TIME_PERIOD"
    }
    headers = {"Accept": "application/vnd.sdmx.genericdata+xml;version=2.1"}
    
    ns = {
        'generic': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/data/generic',
        'message': 'http://www.sdmx.org/resources/sdmxml/schemas/v2_1/message'
    }
    
    try:
        start_time = time.time()
        response = requests.get(DATA_URL, params=params, headers=headers, timeout=REQUEST_TIMEOUT)
        download_time = time.time() - start_time
        
        log(f"Download completed: {len(response.content)/1024/1024:.2f} MB in {download_time:.1f}s")
        
        if response.status_code != 200:
            return {}, [], f"HTTP {response.status_code}"
            
    except requests.exceptions.Timeout:
        return {}, [], "Request timeout"
    except requests.exceptions.RequestException as e:
        return {}, [], str(e)
    
    # Parse XML
    try:
        root = ET.fromstring(response.content)
    except ET.ParseError as e:
        return {}, [], f"XML parse error: {e}"
    
    dataset = root.find('.//message:DataSet', ns)
    if dataset is None:
        return {}, [], "No DataSet found in response"
    
    series_list = dataset.findall('.//generic:Series', ns)
    log(f"Found {len(series_list)} series")
    
    # Extract data: {(territory, product): {period: value}}
    data = defaultdict(dict)
    periods = set()
    territories_found = set()
    products_found = set()
    
    for series in series_list:
        series_key = series.find('.//generic:SeriesKey', ns)
        if series_key is None:
            continue
        
        territory = None
        product = None
        for v in series_key.findall('.//generic:Value', ns):
            vid = v.get('id')
            if vid == 'REF_AREA':
                territory = v.get('value')
                territories_found.add(territory)
            elif vid == 'E_COICOP_REV_ISTAT':
                product = v.get('value')
                products_found.add(product)
        
        if territory is None or product is None:
            continue
        
        for obs in series.findall('.//generic:Obs', ns):
            obs_dim = obs.find('.//generic:ObsDimension', ns)
            obs_value = obs.find('.//generic:ObsValue', ns)
            
            if obs_dim is not None and obs_value is not None:
                period = obs_dim.get('value', '').replace('-', 'M')
                try:
                    data[(territory, product)][period] = float(obs_value.get('value'))
                    periods.add(period)
                except (ValueError, TypeError):
                    pass
    
    log(f"Extracted {len(territories_found)} territories, {len(products_found)} products, {len(data)} combinations")
    return dict(data), sorted(periods), None


def create_excel_file(data: dict, periods: list, product_names: dict, error: str = None) -> io.BytesIO:
    """
    Create Excel file with data and metadata sheets.
    """
    log("Creating Excel file...")
    
    # Create DataFrame
    rows = []
    for (territory, product) in sorted(data.keys()):
        row = {
            'TERRITORY': territory,
            'TERRITORY_NAME': TERRITORY_NAMES.get(territory, territory),
            'PRODUCT_TYPE': product,
            'PRODUCT_NAME': product_names.get(product, product),
        }
        for period in periods:
            row[period] = data[(territory, product)].get(period)
        rows.append(row)
    
    df = pd.DataFrame(rows)
    log(f"DataFrame created: {len(df)} rows x {len(df.columns)} columns")
    
    # Get unique counts
    n_territories = df['TERRITORY'].nunique() if len(df) > 0 else 0
    n_products = df['PRODUCT_TYPE'].nunique() if len(df) > 0 else 0
    
    # Create Excel
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Data sheet
        df.to_excel(writer, sheet_name='Data', index=False)
        
        # Format data sheet
        ws = writer.sheets['Data']
        for col_idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            if col in ['TERRITORY', 'PRODUCT_TYPE']:
                ws.column_dimensions[col_letter].width = 15
            elif col in ['TERRITORY_NAME', 'PRODUCT_NAME']:
                ws.column_dimensions[col_letter].width = 50
            else:
                ws.column_dimensions[col_letter].width = 10
        
        # Metadata sheet
        now = datetime.now()
        metadata = {
            'Field': [
                'edition',
                'edition_type',
                'download_date',
                'source_path',
                'source_path_it',
                'dataflow_id',
                'dataflow_url',
                'measure',
                'measure_code',
                'frequency',
                'frequency_code',
                'base_year',
                'territories',
                'start_period',
                'end_period',
                'n_territories',
                'n_product_types',
                'n_combinations',
                'n_periods',
                'errors'
            ],
            'Value': [
                '',
                'DateDownload',
                now.strftime('%Y-%m-%d %H:%M:%S'),
                SOURCE_PATH,
                SOURCE_PATH_IT,
                DATAFLOW_ID,
                STRUCTURE_URL.split('?')[0],
                'Index numbers',
                '4',
                'Monthly',
                'M',
                '2015',
                'IT (Italia), ITC (Nord-ovest), ITD (Nord-est), ITE (Centro), ITF (Sud), ITG (Isole)',
                periods[0] if periods else '',
                periods[-1] if periods else '',
                n_territories,
                n_products,
                len(df),
                len(periods),
                error if error else 'None'
            ]
        }
        meta_df = pd.DataFrame(metadata)
        meta_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        # Format metadata sheet
        ws_meta = writer.sheets['Metadata']
        ws_meta.column_dimensions['A'].width = 20
        ws_meta.column_dimensions['B'].width = 100
        for row in ws_meta.iter_rows(min_row=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
    
    buffer.seek(0)
    log("Excel file created successfully")
    return buffer


# =============================================================================
# MAIN PIPELINE FUNCTION
# =============================================================================

def run_pipeline() -> dict:
    """
    Main pipeline function.
    Returns dict with status, buffer, and metadata.
    """
    log(f"Pipeline started at {datetime.now().isoformat()}")
    start_time = time.time()
    
    result = {
        'status': 'error',
        'message': '',
        'buffer': None,
        'filename': OUTPUT_FILENAME,
        'metadata': {}
    }
    
    try:
        # 1. Fetch code labels
        product_names = fetch_codelist_names()
        if not product_names:
            log("WARNING: Could not fetch product labels, using codes as names")
        
        # 2. Download data (single request)
        data, periods, error = download_nic_data()
        
        if not data:
            result['message'] = f'Download failed: {error}'
            log(f"ERROR: {result['message']}")
            return result
        
        # 3. Create Excel file
        buffer = create_excel_file(data, periods, product_names, error)
        
        # 4. Get unique counts
        territories = set(t for t, p in data.keys())
        products = set(p for t, p in data.keys())
        
        # 5. Prepare result
        elapsed = time.time() - start_time
        
        result['status'] = 'success'
        result['message'] = f'Downloaded {len(territories)} territories, {len(products)} products, {len(data)} combinations in {elapsed:.1f}s'
        result['buffer'] = buffer
        result['folder_id'] = OUTPUT_FOLDER_ID
        result['metadata'] = {
            'n_territories': len(territories),
            'n_products': len(products),
            'n_combinations': len(data),
            'n_periods': len(periods),
            'period_range': f"{periods[0]} → {periods[-1]}" if periods else '',
            'elapsed_seconds': round(elapsed, 1)
        }
        
        log(f"Pipeline completed successfully in {elapsed:.1f}s")
        
    except Exception as e:
        result['message'] = f'Pipeline error: {str(e)}'
        log(f"ERROR: {result['message']}")
        import traceback
        log(traceback.format_exc())
    
    return result


# =============================================================================
# VERSION CONTROL
# =============================================================================

def get_version_info() -> dict:
    """Return version info for the pipeline."""
    now = datetime.now()
    return {
        'version_type': 'DateDownload',
        'version_value': now.strftime('%YM%m'),
        'check_field': None,
        'archive_name': f"NIC_Tipologia_{now.strftime('%YM%m')}_DateDownload.xlsx"
    }


if __name__ == '__main__':
    result = run_pipeline()
    print(f"\nResult: {result['status']}")
    print(f"Message: {result['message']}")
    if result['metadata']:
        print(f"Metadata: {result['metadata']}")